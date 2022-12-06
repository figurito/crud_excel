from datetime import datetime
from openpyxl import load_workbook





def leer(ruta:str, extraer=str):
    Archivo_excel=load_workbook(ruta)
    Hoja_datos = Archivo_excel['Datos del crud']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]

    info={}
    for i in Hoja_datos:
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{'tarea':i[1].value, 'descripcion':i[2].value,
                                        'estado':i[3].value, 'fecha_inicio':i[4].value,
                                        'fecha_finalizacion':i[5].value})
    
    if not(extraer=='todo'):
        info = filtrar(info, extraer)
    
    for i in info:
        print ('****** Tarea CRUD Excel*******')
        print('Id:'+str(i)+'\n'+'Titulo: '+str(info[i]['tarea'])+'\n'+'Descripcion: '
              +str(info[i]['descripcion'])+'\n'+'Estado:'+str(info[i]['estado'])
              +'\n'+'Fecha Creacion: '+ str(info[i]['fecha_inicio'])
              +'\n'+'Fecha de Finalizacion: ' + str(info[i]['fecha_finalizacion']))
        print()    
    
    return 

def filtrar(info:dict, filtro:str):
    aux={}
    for i in info:
        if info[i]['estado']==filtro:
            aux.setdefault(i,info[i])
    return aux
ruta="C:\\Users\\Acer\\Desktop\\app_consola_crud\\Base_CRUD.xlsx"
#ruta=r"C:\Users\Acer\Desktop\app_consola_crud\Base_CRUD.xlsx"
extraer='todo'
print(leer(ruta, extraer))